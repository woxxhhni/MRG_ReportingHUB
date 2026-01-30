"""
One-off script to create CUSO RAM report Excel template.
Sheets: Model, Issues, MDO_MBO_Rollup (matching sheet_filters / report schema).
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = ROOT / "templates"
OUTPUT = TEMPLATES_DIR / "cuso_ram_report_template.xlsx"

COMMON_SCHEMA = {
    "Model": [
        "ModelID", "ModelName", "ModelOwner", "ModelType",
        "Status", "DateStamp", "RiskLevel", "ValidationDate",
    ],
    "Issues": [
        "IssueID", "IssueType", "Severity", "Status",
        "AssignedTo", "DueDate", "ResolutionDate", "ModelID",
    ],
    "MDO_MBO_Rollup": [
        "RollupID", "RollupName", "ModelCount", "TotalRisk",
        "AverageScore", "LastUpdated", "Owner",
    ],
}


def main():
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for sheet_name, headers in COMMON_SCHEMA.items():
        ws = wb.create_sheet(title=sheet_name)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        ws.column_dimensions["A"].width = 12

    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
