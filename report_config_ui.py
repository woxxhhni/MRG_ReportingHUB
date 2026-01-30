"""
Report Configuration UI
Streamlit application for configuring report parameters
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import date, datetime
from typing import Dict, Any, Optional
import json
import copy

from reports import get_report_manager, list_reports, run_report
from utils.config_manager import (
    get_config_manager,
    get_canonical_config,
    load_report_config,
    save_report_config,
    load_default_config,
    get_default_report_config
)
from utils import MRGDate


# Page configuration
st.set_page_config(
    page_title="MRG Report Configuration",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
if 'config_manager' not in st.session_state:
    st.session_state.config_manager = get_config_manager()

if 'report_manager' not in st.session_state:
    st.session_state.report_manager = get_report_manager()

if 'selected_report' not in st.session_state:
    st.session_state.selected_report = None

if 'current_config' not in st.session_state:
    st.session_state.current_config = None

if 'sheet_columns' not in st.session_state:
    st.session_state.sheet_columns = {}

# Simplified: removed config_source and default_config - only one config per report


def format_report_name_for_display(name: str) -> str:
    """Capitalize the first letter of each word for UI display."""
    return name.title() if name else name


def get_available_reports() -> list:
    """Get list of available reports."""
    return list_reports()


def load_config(report_name: str) -> dict:
    """
    Load configuration for a report.
    Simplified: Only one config file per report.
    
    Returns:
        Configuration dictionary (loaded from file or generated default)
    """
    # Try to load saved config first
    config = load_report_config(report_name)
    if config:
        return config
    
    # If no saved config, try default config file (for initial setup)
    config = load_default_config(report_name)
    if config:
        return config
    
    # Generate new default config
    return get_default_report_config(report_name)


def run_report_with_config(report_name: str, config: dict):
    """
    Run a report with the provided configuration.
    
    Args:
        report_name: Name of the report to run
        config: Configuration dictionary with dates and filters
    """
    from pathlib import Path
    
    try:
        # Get report class
        report_manager = get_report_manager()
        report_class = report_manager.reports.get(report_name.lower())
        
        if not report_class:
            raise ValueError(f"Report '{report_name}' not found")
        
        # Create report instance with config
        # For CUSO RAM Report, pass dates from config
        if "CUSO RAM" in report_name:
            from reports.cuso_ram_report import CUSORAMReport
            inventory_date = config.get("inventory_date")
            compliance_date = config.get("compliance_date")
            report = CUSORAMReport(
                inventory_date=inventory_date,
                compliance_date=compliance_date,
                config_file=None  # Will use config dict we pass
            )
            # Override config with the one from UI
            report.config = config
            report.sheet_filters = config.get("sheet_filters", {})
        else:
            # For other reports, create instance and set config
            report = report_class()
            if hasattr(report, 'config'):
                report.config = config
            if hasattr(report, 'sheet_filters'):
                report.sheet_filters = config.get("sheet_filters", {})
        
        # Run the report
        with st.spinner(f"Running {format_report_name_for_display(report_name)}..."):
            report_path = report.run()
            st.success(f"✅ Report generated successfully!")
            st.info(f"**Report location:** `{report_path}`")
            
            # Show download link if possible
            if isinstance(report_path, Path) and report_path.exists():
                with open(report_path, "rb") as f:
                    st.download_button(
                        label="📥 Download Report",
                        data=f.read(),
                        file_name=report_path.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    except Exception as e:
        st.error(f"Error running report: {e}")
        raise


def get_available_queries(queries_dir: str = "queries") -> list:
    """
    List available .sql query files in the queries folder.
    Returns paths relative to project root, e.g. ['queries/file1.sql', 'queries/file2.sql'].
    """
    root = Path(__file__).resolve().parent
    folder = root / queries_dir
    if not folder.is_dir():
        return []
    paths = sorted(folder.glob("*.sql"), key=lambda p: p.name)
    return [str(p.relative_to(root)).replace("\\", "/") for p in paths]


def get_available_templates(templates_dir: str = "templates") -> list:
    """
    List available .xlsx template files in the templates folder.
    Returns paths relative to project root, e.g. ['templates/report1_template.xlsx', ...].
    """
    root = Path(__file__).resolve().parent
    folder = root / templates_dir
    if not folder.is_dir():
        return []
    paths = sorted(folder.glob("*.xlsx"), key=lambda p: p.name)
    return [str(p.relative_to(root)).replace("\\", "/") for p in paths]


def _common_schema() -> Dict[str, list]:
    """Default sheet/column schema used when creating a new report template."""
    return {
        "Model": [
            "ModelID", "ModelName", "ModelOwner", "ModelType",
            "Status", "DateStamp", "RiskLevel", "ValidationDate",
        ],
        "Issues": [
            "IssueID", "IssueType", "Severity", "Status",
            "AssignedTo", "DueDate", "ResolutionDate", "ModelID",
        ],
        "MDO_MBO_Rollup": [
            "RollupID", "RollupName", "ModelCount", "TotalRisk",
            "AverageScore", "LastUpdated", "Owner",
        ],
    }


def get_sheet_names_from_template(template_path: str) -> list:
    """
    Read sheet names from an Excel template file.
    template_path: path relative to project root or absolute.
    Returns list of sheet names, or [] if file missing/invalid.
    """
    root = Path(__file__).resolve().parent
    path = root / template_path if not Path(template_path).is_absolute() else Path(template_path)
    if not path.exists() or path.suffix.lower() not in (".xlsx", ".xls"):
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def get_sheet_columns_from_template(template_path: str) -> Dict[str, list]:
    """
    Read sheet names and first-row headers from an Excel template.
    Returns Dict[sheet_name, list of column names from row 1].
    """
    root = Path(__file__).resolve().parent
    path = root / template_path if not Path(template_path).is_absolute() else Path(template_path)
    if not path.exists() or path.suffix.lower() not in (".xlsx", ".xls"):
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        result = {}
        for name in wb.sheetnames:
            ws = wb[name]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() if h is not None else "" for h in headers]
            result[name] = [h for h in headers if h]
        wb.close()
        return result
    except Exception:
        return {}


def ensure_report_template(report_name: str, config: Dict[str, Any]) -> str:
    """
    Ensure the report has an Excel template; create a sample if missing.
    Uses config['excel_template_path'] if set and file exists; otherwise creates
    templates/{safe_report_name}_template.xlsx and sets config['excel_template_path'].
    Returns the template path (relative to project root).
    """
    root = Path(__file__).resolve().parent
    templates_dir = root / "templates"
    templates_dir.mkdir(parents=True, exist_ok=True)

    existing = config.get("excel_template_path")
    if existing:
        path = root / existing if not Path(existing).is_absolute() else Path(existing)
        if path.exists():
            return existing

    safe_name = report_name.lower().replace(" ", "_").replace("/", "_")
    template_filename = f"{safe_name}_template.xlsx"
    relative_path = f"templates/{template_filename}"
    out = root / relative_path
    if out.exists():
        config["excel_template_path"] = relative_path
        return relative_path

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    common_schema = _common_schema()
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for sheet_name, headers in common_schema.items():
        ws = wb.create_sheet(title=sheet_name)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        ws.column_dimensions["A"].width = 12
    wb.save(out)
    config["excel_template_path"] = relative_path
    return relative_path


def load_report_schema(report_name: str, config: Dict[str, Any]) -> Dict[str, list]:
    """
    Load schema (sheets and columns) from the report's Excel template.
    If the report has no template or path is invalid, creates a sample template
    and returns schema from it. Schema is used for Tab Queries sheet names and
    Sheet Filter Configuration columns.
    """
    template_path = ensure_report_template(report_name, config)
    schema = get_sheet_columns_from_template(template_path)
    if schema:
        return schema
    return _common_schema()


def get_filter_operators() -> Dict[str, str]:
    """Get available filter operators."""
    return {
        "equals": "Equals (=)",
        "not_equals": "Not Equals (≠)",
        "contains": "Contains",
        "not_contains": "Not Contains",
        "starts_with": "Starts With",
        "ends_with": "Ends With",
        "greater_than": "Greater Than (>)",
        "less_than": "Less Than (<)",
        "greater_equal": "Greater or Equal (≥)",
        "less_equal": "Less or Equal (≤)",
        "in_list": "In List",
        "not_in_list": "Not In List",
        "is_null": "Is Null",
        "is_not_null": "Is Not Null",
        "between": "Between"
    }


def clean_filter_value(value, operator_key: str):
    """
    Clean up corrupted filter values, especially for in_list/not_in_list operators.
    
    Args:
        value: The filter value (could be corrupted)
        operator_key: The operator type
        
    Returns:
        Cleaned value in the correct format
    """
    if operator_key in ["in_list", "not_in_list", "contains", "not_contains"]:
        if isinstance(value, list):
            cleaned_list = []
            for item in value:
                if isinstance(item, str):
                    item_clean = item.strip()
                    # Remove leading/trailing brackets, quotes, and whitespace
                    item_clean = item_clean.strip('[]"\'')
                    # Try to detect if this is a corrupted serialized list
                    if item_clean.startswith('[') or (item_clean.startswith("'") and item_clean.endswith("'")):
                        try:
                            import ast
                            # Try to safely evaluate the string
                            parsed = ast.literal_eval(item_clean)
                            if isinstance(parsed, list):
                                cleaned_list.extend([str(v).strip('"\'') for v in parsed])
                            else:
                                cleaned_list.append(str(parsed).strip('"\''))
                        except:
                            # If parsing fails, just clean the string
                            cleaned_list.append(item_clean.strip('"\''))
                    else:
                        cleaned_list.append(item_clean)
            return cleaned_list
        elif isinstance(value, str):
            # If it's a string, try to parse it as comma-separated or JSON
            try:
                import ast
                parsed = ast.literal_eval(value)
                if isinstance(parsed, list):
                    return [str(v).strip('"\'') for v in parsed]
            except:
                pass
            # If parsing fails, treat as comma-separated string
            return [v.strip().strip('"\'') for v in value.split(",") if v.strip()]
        return []
    elif operator_key == "between":
        if isinstance(value, list) and len(value) >= 2:
            return [str(value[0]).strip(), str(value[1]).strip()]
        return []
    else:
        # For other operators, return as string
        if isinstance(value, list):
            # If somehow a list got saved for a non-list operator, take first element
            return str(value[0]) if value else ""
        return str(value) if value else ""

    return value


def render_column_filter(sheet_name: str, column_name: str, 
                        current_filter: Optional[Dict] = None) -> Dict[str, Any]:
    """
    Render UI for a single column filter.
    
    Returns:
        Filter configuration dictionary or None if filter is disabled
    """
    col1, col2, col3 = st.columns([2, 2, 1])
    
    with col1:
        operator_key = st.selectbox(
            "Operator",
            options=list(get_filter_operators().keys()),
            format_func=lambda x: get_filter_operators()[x],
            key=f"filter_{sheet_name}_{column_name}_operator",
            index=list(get_filter_operators().keys()).index(current_filter.get("operator", "equals")) if current_filter and "operator" in current_filter else 0
        )
    
    with col2:
        if operator_key in ["is_null", "is_not_null"]:
            # No value needed for null checks
            value = ""
        elif operator_key in ["in_list", "not_in_list"]:
            # Multi-value input
            # Clean the current value first
            current_value = current_filter.get("value", []) if current_filter else []
            cleaned_value = clean_filter_value(current_value, operator_key)
            
            # Convert cleaned list to comma-separated string for display (no brackets)
            if isinstance(cleaned_value, list):
                # Join list items with comma and space, no brackets
                value_str_default = ", ".join(str(v).strip('[]"\'') for v in cleaned_value if v) if cleaned_value else ""
            else:
                # If it's a string, try to remove any brackets if present
                value_str = str(cleaned_value) if cleaned_value else ""
                # Remove brackets if the string looks like a list representation
                if value_str.strip().startswith('[') and value_str.strip().endswith(']'):
                    # Try to parse and extract values
                    try:
                        import ast
                        parsed = ast.literal_eval(value_str)
                        if isinstance(parsed, list):
                            value_str_default = ", ".join(str(v).strip('[]"\'') for v in parsed if v)
                        else:
                            value_str_default = value_str.strip('[]"\'')
                    except:
                        # If parsing fails, just remove brackets manually
                        value_str_default = value_str.strip('[]"\'')
                else:
                    value_str_default = value_str
            
            value_str = st.text_input(
                "Value (comma-separated)",
                value=value_str_default,
                key=f"filter_{sheet_name}_{column_name}_value"
            )
            # Convert string back to list, cleaning up any extra whitespace and brackets
            # Remove any brackets that might be in the input string
            cleaned_input = value_str.strip().strip('[]')
            # Split by comma and clean each value (remove brackets, quotes, whitespace)
            value = [v.strip().strip('[]"\'') for v in cleaned_input.split(",") if v.strip()] if cleaned_input else []
        elif operator_key == "between":
            # Range input
            # Handle both list and other formats
            current_value = current_filter.get("value", []) if current_filter else []
            if isinstance(current_value, list) and len(current_value) >= 2:
                value_min_default = str(current_value[0]) if current_value[0] else ""
                value_max_default = str(current_value[1]) if current_value[1] else ""
            else:
                value_min_default = ""
                value_max_default = ""
            
            col2a, col2b = st.columns(2)
            with col2a:
                value_min = st.text_input(
                    "Min Value",
                    value=value_min_default,
                    key=f"filter_{sheet_name}_{column_name}_value_min"
                )
            with col2b:
                value_max = st.text_input(
                    "Max Value",
                    value=value_max_default,
                    key=f"filter_{sheet_name}_{column_name}_value_max"
                )
            value = [value_min, value_max] if value_min or value_max else []
        else:
            # Single value input
            # Clean the current value first (handle cases where user entered list format)
            current_value = current_filter.get("value", "") if current_filter else ""
            cleaned_value = clean_filter_value(current_value, operator_key)
            
            # For single value operators, ensure we have a string (no brackets)
            if isinstance(cleaned_value, list):
                # If user entered list format, join them without brackets
                value_default = ", ".join(str(v).strip('[]"\'') for v in cleaned_value if v) if cleaned_value else ""
            else:
                # Remove brackets if present in string
                value_str = str(cleaned_value) if cleaned_value else ""
                if value_str.strip().startswith('[') and value_str.strip().endswith(']'):
                    value_default = value_str.strip('[]"\'')
                else:
                    value_default = value_str
            
            value = st.text_input(
                "Value",
                value=value_default,
                key=f"filter_{sheet_name}_{column_name}_value"
            )
    
    with col3:
        if st.button("Delete", key=f"delete_{sheet_name}_{column_name}"):
            return None

    # Save contains/not_contains as list in config (consistent with in_list)
    if operator_key in ["contains", "not_contains"] and isinstance(value, str):
        value = [v.strip() for v in value.split(",") if v.strip()]

    # Always return filter config with enabled=True (all displayed filters are enabled)
    return {
        "operator": operator_key,
        "value": value,
        "enabled": True
    }


def main():
    """Main application."""
    st.title("📊 MRG Report Configuration")
    st.markdown("---")

    # Sidebar - Report List
    with st.sidebar:
        st.header("📋 Reports")
        
        available_reports = get_available_reports()
        
        if not available_reports:
            st.warning("No reports available. Please register reports first.")
            return
        
        # Report selection (display report names in title case)
        selected_report = st.selectbox(
            "Select Report",
            options=available_reports,
            format_func=format_report_name_for_display,
            key="report_selector"
        )
        
        # Only reload config if report actually changed
        current_selected = st.session_state.get('selected_report')
        if selected_report != current_selected:
            st.session_state.selected_report = selected_report
            try:
                config = load_config(selected_report)
                # Schema from report's Excel template (creates sample template if missing)
                st.session_state.sheet_columns = load_report_schema(selected_report, config)
                st.session_state.current_config = copy.deepcopy(config)
            except Exception as e:
                st.error(f"Error loading report configuration: {e}")
                fallback_config = get_default_report_config(selected_report)
                st.session_state.sheet_columns = load_report_schema(selected_report, fallback_config)
                st.session_state.current_config = copy.deepcopy(fallback_config)
        
        st.markdown("---")
        
        # Configuration actions
        st.subheader("Configuration Actions")
        
        # Save and Run buttons (equal width and height)
        # Add CSS to ensure buttons have the same height
        st.markdown("""
        <style>
        .stButton > button {
            height: 3rem;
        }
        </style>
        """, unsafe_allow_html=True)
        
        action_col1, action_col2 = st.columns([1, 1])
        with action_col1:
            if st.button("💾 Save Configuration", use_container_width=True, type="primary"):
                if st.session_state.current_config:
                    try:
                        save_path = save_report_config(selected_report, st.session_state.current_config)
                        st.success(f"✅ Configuration saved to: `{save_path.name}`")
                    except Exception as e:
                        st.error(f"Failed to save configuration: {e}")
        
        with action_col2:
            if st.button("🚀 Run Report", use_container_width=True, type="primary"):
                if st.session_state.current_config:
                    try:
                        # Run report with current configuration
                        run_report_with_config(selected_report, st.session_state.current_config)
                    except Exception as e:
                        st.error(f"Failed to run report: {e}")
                        st.exception(e)
        
        # Reset button
        if st.button("🔄 Reset to Default", use_container_width=True):
            try:
                # Load default config (from default file or generate)
                default_config = load_default_config(selected_report)
                if default_config is None:
                    default_config = get_default_report_config(selected_report)
                st.session_state.current_config = copy.deepcopy(default_config)
                st.success("✅ Reset to default configuration!")
            except Exception as e:
                st.error(f"Error resetting to default: {e}")
        
        st.markdown("---")
        
        # Show current config file path
        config_path = st.session_state.config_manager.get_config_path(selected_report, "config")
        st.caption(f"**Config file:** `{config_path}`")
    
    # Main content area
    if st.session_state.selected_report:
        report_name = st.session_state.selected_report
        config = st.session_state.current_config
        sheet_columns = st.session_state.sheet_columns or {}

        # Top of page: report name and description
        st.header(f"Configuration: {format_report_name_for_display(report_name)}")
        config.setdefault("report_description", "")
        desc = config.get("report_description") or ""
        if not isinstance(desc, str):
            desc = str(desc)
        report_description = st.text_area(
            "Report Description",
            value=desc,
            height=100,
            placeholder="Enter a short description of this report (saved in configuration).",
            key="report_description_input"
        )
        config["report_description"] = (report_description or "").strip()

        st.markdown("---")

        # Report template: display path and allow selecting from templates folder
        config.setdefault("excel_template_path", None)
        available_templates = get_available_templates()
        current_path = config.get("excel_template_path") or ""
        st.subheader("📄 Report Template")
        template_options = ["(No template)"] + list(available_templates)
        if current_path and current_path not in template_options:
            template_options = [current_path] + template_options
        idx = template_options.index(current_path) if current_path in template_options else 0
        selected_template = st.selectbox(
            "Excel template",
            options=template_options,
            index=idx,
            format_func=lambda x: x if x and x != "(No template)" else "(No template)",
            key="excel_template_selector"
        )
        if selected_template and selected_template != "(No template)":
            config["excel_template_path"] = selected_template
            if selected_template != current_path:
                new_schema = get_sheet_columns_from_template(selected_template)
                if new_schema:
                    st.session_state.sheet_columns = new_schema
                    sheet_columns = new_schema
        else:
            config["excel_template_path"] = None
            st.session_state.sheet_columns = {}
        st.caption(f"**Current template:** `{config.get('excel_template_path') or '(none)'}`")

        # Refresh sheet_columns after possible template change
        sheet_columns = st.session_state.sheet_columns or {}

        # Align config with template: sheet_filters and tab_queries only for sheets in this report's template
        template_sheet_names = list(sheet_columns.keys())
        if "sheet_filters" not in config:
            config["sheet_filters"] = {}
        # Keep only filters for sheets in template; only keep column filters that exist in template
        allowed_columns = {name: set(sheet_columns.get(name, [])) for name in template_sheet_names}
        old_sheet_filters = config.get("sheet_filters") if isinstance(config.get("sheet_filters"), dict) else {}
        config["sheet_filters"] = {}
        for name in template_sheet_names:
            existing = old_sheet_filters.get(name) if isinstance(old_sheet_filters.get(name), dict) else {}
            config["sheet_filters"][name] = {
                col: val for col, val in existing.items()
                if col in allowed_columns.get(name, set())
            }
        if "tab_queries" not in config:
            config["tab_queries"] = {}
        tab_queries_raw = config.get("tab_queries") or {}
        if not isinstance(tab_queries_raw, dict):
            tab_queries_raw = {}
        config["tab_queries"] = {k: v for k, v in tab_queries_raw.items() if k in template_sheet_names}

        st.markdown("---")

        # Date Configuration
        st.subheader("📅 Date Configuration")
        date_col1, date_col2 = st.columns(2)
        
        # Safely parse inventory date
        try:
            inventory_date_str = config.get("inventory_date", date.today().strftime("%Y-%m-%d"))
            if isinstance(inventory_date_str, str):
                inventory_date_value = datetime.strptime(inventory_date_str, "%Y-%m-%d").date()
            else:
                inventory_date_value = date.today()
        except (ValueError, TypeError):
            inventory_date_value = date.today()
        
        with date_col1:
            inventory_date = st.date_input(
                "Inventory Date",
                value=inventory_date_value,
                key="inventory_date_input"
            )
            config["inventory_date"] = inventory_date.strftime("%Y-%m-%d")
        
        # Safely parse compliance date
        try:
            compliance_date_str = config.get("compliance_date", date.today().strftime("%Y-%m-%d"))
            if isinstance(compliance_date_str, str):
                compliance_date_value = datetime.strptime(compliance_date_str, "%Y-%m-%d").date()
            else:
                compliance_date_value = date.today()
        except (ValueError, TypeError):
            compliance_date_value = date.today()
        
        with date_col2:
            compliance_date = st.date_input(
                "Compliance Date",
                value=compliance_date_value,
                key="compliance_date_input"
            )
            config["compliance_date"] = compliance_date.strftime("%Y-%m-%d")

        # Tab queries: select query file per sheet (sheets from report schema)
        config.setdefault("excel_template_path", None)
        config.setdefault("query", None)
        config.setdefault("tab_queries", {})

        tab_queries = config.get("tab_queries") or {}
        if not isinstance(tab_queries, dict):
            tab_queries = {}
        available_queries = get_available_queries()
        # Use template sheet order (same as in template file)
        sheet_names = template_sheet_names

        st.markdown("---")
        st.subheader("📁 Tab Queries")
        st.caption("Assign a query file to each output sheet. Sheets and order come from this report's Excel template.")

        if not sheet_names:
            st.info("No sheet schema loaded. Select a report and ensure sheet schema is available (e.g. Model, Issues, MDO_MBO_Rollup).")
        elif not available_queries:
            st.warning("No .sql files found in the queries folder. Add query files to enable tab queries.")
        else:
            for sheet_name in sheet_names:
                current = tab_queries.get(sheet_name) or "(None)"
                options = ["(None)"]
                if current and current != "(None)" and current not in available_queries:
                    options.append(current)
                options.extend(available_queries)
                idx = options.index(current) if current in options else 0
                selected = st.selectbox(
                    f"**{sheet_name}**",
                    options=options,
                    index=idx,
                    key=f"tab_query_{sheet_name}"
                )
                if selected and selected != "(None)":
                    tab_queries[sheet_name] = selected
                elif sheet_name in tab_queries:
                    del tab_queries[sheet_name]
        config["tab_queries"] = tab_queries

        st.markdown("---")

        # Sheet Filter Configuration (sheets and columns from this report's template)
        st.subheader("🔍 Sheet Filter Configuration")
        st.caption("Filters for each sheet and column. Sheets and columns come from this report's Excel template.")

        # Reset All Filters: clear filters but keep template sheet structure
        filter_action_col1, filter_action_col2 = st.columns([1, 4])
        with filter_action_col1:
            if st.button("🔄 Reset All Filters", use_container_width=True, type="secondary"):
                config["sheet_filters"] = {name: {} for name in template_sheet_names}
                st.success("✅ All filters have been reset!")

        if not sheet_columns:
            st.info("No sheet schema available. Select an Excel template above to define sheets and columns for this report.")
        else:
            # Display filters for each sheet
            for sheet_name, columns in sheet_columns.items():
                with st.expander(f"📄 Sheet: **{sheet_name}** ({len(columns)} columns)", expanded=True):
                    if sheet_name not in config["sheet_filters"]:
                        config["sheet_filters"][sheet_name] = {}
                    
                    if not columns:
                        st.info(f"No columns defined for sheet '{sheet_name}'.")
                    else:
                        # Column selection
                        selected_columns = st.multiselect(
                            f"Select columns to filter",
                            options=columns,
                            default=list(config["sheet_filters"][sheet_name].keys()),
                            key=f"column_selector_{sheet_name}"
                        )
                        
                        # Display filters for selected columns
                        for column_name in selected_columns:
                            st.markdown(f"**Column: {column_name}**")
                            
                            current_filter = config["sheet_filters"][sheet_name].get(column_name)
                            
                            # Clean the filter value before rendering (fix corrupted values)
                            if current_filter and "operator" in current_filter:
                                current_filter = current_filter.copy()  # Don't modify original
                                current_filter["value"] = clean_filter_value(
                                    current_filter.get("value"), 
                                    current_filter.get("operator", "")
                                )
                            
                            filter_config = render_column_filter(sheet_name, column_name, current_filter)
                            
                            # filter_config is None only if user clicked "Delete" button
                            # Otherwise, it always returns a config dict with enabled=True
                            # All displayed filters are automatically enabled
                            if filter_config is not None:
                                config["sheet_filters"][sheet_name][column_name] = filter_config
                            elif column_name in config["sheet_filters"][sheet_name]:
                                # Only remove if user explicitly clicked delete button
                                del config["sheet_filters"][sheet_name][column_name]
                        
                        # Remove filters for unselected columns
                        columns_to_remove = [
                            col for col in config["sheet_filters"][sheet_name].keys()
                            if col not in selected_columns
                        ]
                        for col in columns_to_remove:
                            del config["sheet_filters"][sheet_name][col]
        
        # Update session state
        st.session_state.current_config = config
        
        st.markdown("---")
        
        # Configuration Preview (canonical format, same as saved)
        canonical_config = get_canonical_config(config)
        with st.expander("📋 Configuration Preview (JSON)"):
            st.json(canonical_config)

        # Download configuration (same format as saved config)
        st.download_button(
            label="📥 Download Configuration as JSON",
            data=json.dumps(canonical_config, indent=2, ensure_ascii=False),
            file_name=f"{report_name.lower().replace(' ', '_')}_config.json",
            mime="application/json"
        )
    else:
        st.info("Please select a report from the sidebar to configure.")


if __name__ == "__main__":
    main()
