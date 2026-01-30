"""
Excel Manager
Handles Excel file operations including saving DataFrames to xlsx files
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Optional, List, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import logging

logger = logging.getLogger(__name__)


class ExcelManager:
    """
    Manager for Excel file operations.
    Handles saving DataFrames to Excel files with formatting options.
    """
    
    def __init__(self, output_dir: Union[str, Path] = "output_data"):
        """
        Initialize Excel Manager.
        
        Args:
            output_dir: Base directory for output Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Excel Manager initialized. Output directory: {self.output_dir}")
    
    def save_dataframe(self, df: pd.DataFrame, file_path: Union[str, Path],
                      sheet_name: str = "Sheet1", index: bool = False,
                      startrow: Optional[int] = None, startcol: Optional[int] = None,
                      start_cell: Optional[str] = None,
                      format_header: bool = True, auto_adjust_width: bool = True) -> Path:
        """
        Save a single DataFrame to an Excel file.
        
        Args:
            df: pandas DataFrame to save
            file_path: Path to Excel file (str or Path)
            sheet_name: Name of the sheet (default: "Sheet1")
            index: Whether to include DataFrame index (default: False)
            startrow: Starting row position (1-based, default: 1 if start_cell not provided)
            startcol: Starting column position (1-based, default: 1 if start_cell not provided)
            start_cell: Excel cell reference like "A1", "B3" (takes precedence over startrow/startcol)
            format_header: Whether to format header row (default: True)
            auto_adjust_width: Whether to auto-adjust column widths (default: True)
            
        Returns:
            Path to saved Excel file
            
        Examples:
            # Using Excel cell reference
            excel_mgr.save_dataframe(df, "report.xlsx", start_cell="A1")
            excel_mgr.save_dataframe(df, "report.xlsx", start_cell="B3")
            
            # Using row/column numbers
            excel_mgr.save_dataframe(df, "report.xlsx", startrow=1, startcol=1)
        """
        excel_path = Path(file_path)
        
        # Ensure .xlsx extension
        if excel_path.suffix.lower() != '.xlsx':
            excel_path = excel_path.with_suffix('.xlsx')
        
        # Create parent directories if needed
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Parse start_cell if provided, otherwise use startrow/startcol
        if start_cell:
            startrow, startcol = self._parse_cell_reference(start_cell)
        else:
            # Default to 1-based indexing (Excel standard)
            startrow = startrow if startrow is not None else 1
            startcol = startcol if startcol is not None else 1
        
        try:
            # If file exists, load it; otherwise create new workbook
            if excel_path.exists():
                wb = load_workbook(excel_path)
                # Remove sheet if it already exists
                if sheet_name in wb.sheetnames:
                    wb.remove(wb[sheet_name])
                ws = wb.create_sheet(title=sheet_name)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
            
            # Write DataFrame to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=True), startrow):
                for c_idx, value in enumerate(row, startcol):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Format header row
            if format_header and len(df) > 0:
                self._format_header_row(ws, startrow, startcol - 1, len(df.columns))
            
            # Auto-adjust column widths
            if auto_adjust_width:
                self._auto_adjust_column_widths(ws, startcol - 1, len(df.columns))
            
            # Save workbook
            wb.save(excel_path)
            start_info = start_cell if start_cell else f"Row{startrow}Col{startcol}"
            logger.info(f"DataFrame saved to Excel: {excel_path} (Sheet: {sheet_name}, Rows: {len(df)}, Start: {start_info})")
            
            return excel_path
        
        except Exception as e:
            logger.error(f"Failed to save DataFrame to Excel: {e}")
            raise
    
    def save_multiple_dataframes(self, dataframes: Dict[str, pd.DataFrame],
                                file_path: Union[str, Path], index: bool = False,
                                format_header: bool = True, auto_adjust_width: bool = True) -> Path:
        """
        Save multiple DataFrames to a single Excel file, each in its own sheet.
        
        Args:
            dataframes: Dictionary mapping sheet names to DataFrames
            file_path: Path to Excel file (str or Path)
            index: Whether to include DataFrame index (default: False)
            format_header: Whether to format header rows (default: True)
            auto_adjust_width: Whether to auto-adjust column widths (default: True)
            
        Returns:
            Path to saved Excel file
            
        Example:
            dataframes = {
                "Summary": df_summary,
                "Detail": df_detail,
                "Trends": df_trends
            }
            excel_manager.save_multiple_dataframes(dataframes, "report.xlsx")
        """
        excel_path = Path(file_path)
        
        # Ensure .xlsx extension
        if excel_path.suffix.lower() != '.xlsx':
            excel_path = excel_path.with_suffix('.xlsx')
        
        # Create parent directories if needed
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            # Create new workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add each DataFrame as a new sheet
            for sheet_name, df in dataframes.items():
                if df.empty:
                    logger.warning(f"DataFrame for sheet '{sheet_name}' is empty, skipping")
                    continue
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Write DataFrame to worksheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format header row
                if format_header:
                    self._format_header_row(ws, 1, 0, len(df.columns))
                
                # Auto-adjust column widths
                if auto_adjust_width:
                    self._auto_adjust_column_widths(ws, 0, len(df.columns))
                
                logger.info(f"Added sheet '{sheet_name}' with {len(df)} rows")
            
            # Save workbook
            wb.save(excel_path)
            total_sheets = len(dataframes)
            logger.info(f"Saved {total_sheets} DataFrames to Excel: {excel_path}")
            
            return excel_path
        
        except Exception as e:
            logger.error(f"Failed to save multiple DataFrames to Excel: {e}")
            raise

    def save_dataframes_to_template(
        self,
        template_path: Union[str, Path],
        dataframes: Dict[str, pd.DataFrame],
        output_path: Union[str, Path],
        index: bool = False,
        format_header: bool = True,
        auto_adjust_width: bool = True,
    ) -> Path:
        """
        Load an Excel template and write DataFrames to named sheets.
        Sheets are created if they do not exist; existing sheet content is overwritten from row 1.

        Args:
            template_path: Path to the Excel template file
            dataframes: Dict mapping sheet name to DataFrame
            output_path: Path for the output Excel file
            index: Whether to include DataFrame index (default: False)
            format_header: Whether to format header row (default: True)
            auto_adjust_width: Whether to auto-adjust column widths (default: True)

        Returns:
            Path to the saved output file
        """
        template_path = Path(template_path)
        output_path = Path(output_path)
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        try:
            wb = load_workbook(template_path)
            for sheet_name, df in dataframes.items():
                if df.empty:
                    logger.warning(f"DataFrame for sheet '{sheet_name}' is empty, skipping")
                    continue
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
                # Clear existing content (from row 1) and write DataFrame
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.value = None
                for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                if format_header:
                    self._format_header_row(ws, 1, 0, len(df.columns))
                if auto_adjust_width:
                    self._auto_adjust_column_widths(ws, 0, len(df.columns))
                logger.info(f"Wrote sheet '{sheet_name}' with {len(df)} rows")
            wb.save(output_path)
            logger.info(f"Saved report to template output: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Failed to save DataFrames to template: {e}")
            raise

    def append_dataframe(self, df: pd.DataFrame, file_path: Union[str, Path],
                        sheet_name: str = "Sheet1", index: bool = False,
                        start_cell: Optional[str] = None,
                        format_header: bool = True, auto_adjust_width: bool = True) -> Path:
        """
        Append a DataFrame to an existing Excel file (adds to end of existing data).
        
        Args:
            df: pandas DataFrame to append
            file_path: Path to existing Excel file
            sheet_name: Name of the sheet to append to
            index: Whether to include DataFrame index
            start_cell: Optional Excel cell reference like "A1" to specify start position
                       (if not provided, appends to end of existing data)
            format_header: Whether to format header row (only if sheet is new)
            auto_adjust_width: Whether to auto-adjust column widths
            
        Returns:
            Path to Excel file
        """
        excel_path = Path(file_path)
        
        if not excel_path.exists():
            logger.warning(f"File {excel_path} does not exist. Creating new file.")
            return self.save_dataframe(df, excel_path, sheet_name, index=index,
                                     start_cell=start_cell,
                                     format_header=format_header, 
                                     auto_adjust_width=auto_adjust_width)
        
        try:
            wb = load_workbook(excel_path)
            
            # Get or create worksheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Use start_cell if provided, otherwise find next empty row
                if start_cell:
                    startrow, startcol = self._parse_cell_reference(start_cell)
                else:
                    startrow = ws.max_row + 1
                    startcol = 1
                # Don't write header if appending to existing sheet
                write_header = False
            else:
                ws = wb.create_sheet(title=sheet_name)
                if start_cell:
                    startrow, startcol = self._parse_cell_reference(start_cell)
                else:
                    startrow = 1
                    startcol = 1
                write_header = True
            
            # Write DataFrame
            for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=write_header), startrow):
                for c_idx, value in enumerate(row, startcol):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Format header if this is a new sheet
            if write_header and format_header:
                self._format_header_row(ws, startrow, startcol - 1, len(df.columns))
            
            # Auto-adjust column widths
            if auto_adjust_width:
                self._auto_adjust_column_widths(ws, startcol - 1, len(df.columns))
            
            wb.save(excel_path)
            start_info = start_cell if start_cell else f"Row{startrow}Col{startcol}"
            logger.info(f"Appended {len(df)} rows to sheet '{sheet_name}' in {excel_path} (Start: {start_info})")
            
            return excel_path
        
        except Exception as e:
            logger.error(f"Failed to append DataFrame to Excel: {e}")
            raise
    
    def read_excel(self, file_path: Union[str, Path], sheet_name: Optional[str] = None,
                  **kwargs) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        Read Excel file and return DataFrame(s).
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name to read (None reads all sheets)
            **kwargs: Additional arguments passed to pd.read_excel
            
        Returns:
            DataFrame if sheet_name specified, or dict of DataFrames if None
        """
        excel_path = Path(file_path)
        
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        try:
            if sheet_name:
                df = pd.read_excel(excel_path, sheet_name=sheet_name, **kwargs)
                logger.info(f"Read sheet '{sheet_name}' from {excel_path} ({len(df)} rows)")
                return df
            else:
                # Read all sheets
                dfs = pd.read_excel(excel_path, sheet_name=None, **kwargs)
                logger.info(f"Read {len(dfs)} sheets from {excel_path}")
                return dfs
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise
    
    def _parse_cell_reference(self, cell_ref: str) -> tuple[int, int]:
        """
        Parse Excel cell reference (e.g., "A1", "B3") to row and column numbers.
        
        Args:
            cell_ref: Excel cell reference like "A1", "B3", "AA10"
            
        Returns:
            Tuple of (row_number, column_number) both 1-based
        """
        # Match pattern like "A1", "B3", "AA10", etc.
        pattern = r'^([A-Z]+)(\d+)$'
        match = re.match(pattern, cell_ref.upper())
        
        if not match:
            raise ValueError(f"Invalid cell reference format: {cell_ref}. Expected format like 'A1', 'B3'")
        
        col_letter = match.group(1)
        row_number = int(match.group(2))
        col_number = column_index_from_string(col_letter)
        
        return row_number, col_number
    
    def _format_header_row(self, ws, row_num: int, start_col: int, num_cols: int):
        """Format header row with styling."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx in range(start_col + 1, start_col + num_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
    
    def _auto_adjust_column_widths(self, ws, start_col: int, num_cols: int):
        """Auto-adjust column widths based on content."""
        for col_idx in range(start_col + 1, start_col + num_cols + 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with some padding, but cap at 50
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def list_sheets(self, file_path: Union[str, Path]) -> List[str]:
        """
        List all sheet names in an Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        excel_path = Path(file_path)
        
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        try:
            wb = load_workbook(excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            logger.error(f"Failed to list sheets: {e}")
            raise
